import openpyxl

#获取表格
wb = openpyxl.load_workbook('temperature_field.xlsx')
#print(wb)

# 获取所有表单
#sheets = wb.worksheets
#print(sheets)

# 通过索引获取表单，索引从 0 开始
#sheet = wb.worksheets[0]
#print(sheet)

# 通过表单名字获取表单.有个问题，之后使用 sheet 后无法 点 出提示
#sheet = wb.get_sheet_by_name("temperature_field")
#print(sheet)

# 简洁版 通过字典的形式. 问题和上面一样
sheet = wb['temperature_field']
print(sheet)

# 获取一行的数据
#cell = sheet[1]
#print(cell)

#  输出数据  (<Cell 'Sheet1'.A1>, <Cell 'Sheet1'.B1>, <Cell 'Sheet1'.C1>)

# 获取一列的数据
#cell1 = sheet['A']
#print(cell1)

# 获取所有的行，最简单的方法
#data = sheet.rows
#data_list = list(data)[1:]  # 剔除第一条数据，第一条是标题
#print(data_list)

# 获取所有数据
data_list = list(sheet.rows)[1:]  # 剔除第一条数据，第一条是标题
data = []  # 新列表储存数据#
for row in data_list:
    # 每一行的数据存储起来 每循环一次清空
    row_data = []
    # 把 row 里面的数据依次提取出来
    for cell in row:
        # 把值添加到 row_data 列表中
        row_data.append(cell.value)
    # 每循环一次，往 new_data 添加一次数据
    data.append(row_data)
print(data)
wb.close()

#按序号分组
DATA = [[] for i in range(100)]   #先分100列
for x in data:
    y = int(x[2])
    DATA[y-1].append(x)
print(DATA)

#新建新的表格
#使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
newwb = openpyxl.Workbook()

# 默认sheet
shet = newwb.active
shet.title = "temperature_field"

#ws1 = wb.create_sheet()   #默认插在工作簿末尾
#ws2 = wb.create_sheet(0)  #插入在工作簿的第一个位置


#将数据按对应位置赋值到单元格内
for j in range(len(DATA)):
    for i in range(len(DATA[j])):
        zi = ''
        zi += "{: <5}{: <2}{: <3}{: <6}{: <6}{: <3}{: <4}{: <3}".format(str(DATA[j][i][0]),str(DATA[j][i][1]),str(DATA[j][i][2]),str(DATA[j][i][3]),str(DATA[j][i][4]),str(DATA[j][i][6]),str(DATA[j][i][7]),str(DATA[j][i][8]))
        zi += DATA[j][i][5].strftime('%Y/%m/%d %H:%M:%S')
        shet.cell(i+1, j+1).value=zi
        print(zi)

#保存
newwb.save('newexl.xlsx')
print('successful convert!')

