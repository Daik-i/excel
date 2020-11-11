#将txt数据导入成一个列表，每一行数据是一个子列表
datas = []
with open("input/dataD.txt") as fin:
    for line in fin:
        line = line[:-1]
        datas.append(line.split("\t"))
#清空txt
a = open("input/dataD.txt", "w")
a.write("")
a.close()

#引入所需库
# pip install openpyxl
import openpyxl
from datetime import datetime

#为页的名称加上时间
time = datetime.now()
now = str(time)[8:10]+'-'+str(time)[11:13]+' '+str(time)[14:16]

#打开excel
workbook = openpyxl.load_workbook("output/dataD.xlsx")

# 默认sheet
# sheet = workbook.active
# sheet.title = "默认sheet"

# 创建新页并设定名称
sheet = workbook.create_sheet(title=now)

#数据整理，把并列的数据拆分
newdata = []
for shuju in datas:
    if len(shuju[0]) > 39:  #用[0]将子列表的字符串取出
        n = 39  # 大列表中几个数据组成一个小列表
        for i in range(0, len(shuju[0]), n):
            newdata.append([shuju[0][i:i + n]])
    else:
        newdata.append(shuju)

#将16进制字符转换成对应数字
def zhuanhuan1(z):
    if z in 'A':
        num1 = 10
    elif z in 'B':
        num1 = 11
    elif z in 'C':
        num1 = 12
    elif z in 'D':
        num1 = 13
    elif z in 'E':
        num1 = 14
    elif z in 'F':
        num1 = 15
    else:
        num1 = int(z)
    return num1

def zhuanhuan(A):
    num = zhuanhuan1(A[0])*16+zhuanhuan1(A[1])
    return num

#将数据调整为正常数据,部分由字符串变为数字
for i in newdata:
    xuhao = str(zhuanhuan(i[0][9:11]))
    wendu = round((zhuanhuan(i[0][12:14]) + zhuanhuan(i[0][15:17]) * 256) * 0.00268127 - 46.85, 2)
    shidu = round((zhuanhuan(i[0][18:20]) + zhuanhuan(i[0][21:23]) * 256) * 0.00190735 - 6.00, 2)
    dianya = zhuanhuan(i[0][24:26])
    jiange = int(i[0][27:29])*10
    #gonglv = int(i[0][30:32])
    #geshu = int(i[0][33:35])
    i[0] = i[0][:8]+' '+str(xuhao).zfill(2)+' '+str(wendu).center(5,' ')+' '+str(shidu).center(5,' ')+' '+str(dianya)+' '+str(jiange).center(3,' ')+' '+i[0][30:]


#将数据分到对应子列表集
DATA = [[] for i in range(19)]
for x in newdata:
    y = int(x[0][9:11])
    if 54 < y < 73:
        DATA[y-1-54].append(x)
    else:
        DATA[18].append(x)

#将数据按对应位置赋值到单元格内
for j in range(len(DATA)):
    for i in range(len(DATA[j])):
        sheet.cell(i+1, j+1).value=DATA[j][i][0]

#保存
workbook.save('output/dataD.xlsx')
print('successful convert!')