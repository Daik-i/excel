

# !/usr/bin/python3
import time
import pymysql
import openpyxl
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection

# 打开数据库连接
while True:
    try:
        db = pymysql.connect("raspberrypi", "root", "123", "data1")
        break;
    except Exception as e:
        print(str(e))
        time.sleep(2)



# 使用 cursor() 方法创建一个游标对象 cursor
cursor = db.cursor()

# 使用 execute()  方法执行 SQL 查询
# cursor.execute("SELECT VERSION()")

# Select * from emp where serial_number between 1500 and 3000;
# 查询 emp 表中 SAL 列中大于 1500 的小于 3000 的值。
# 注意：大于等于 1500 且小于等于 3000， 1500 为下限，3000 为上限，下限在前，上限在后，查询的范围包涵有上下限的值。

# # 当前时间
# now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
#
# datetime1 = datetime.datetime(2020,8,18,1,22,0)
# datetime2 = datetime.datetime(2020,8,20,1,22,0)

sql_str_read = '''select * from temperature_field WHERE  submission_date between '2020-08-20 20:15:37' and NOW();'''



cursor.execute(sql_str_read)
dataFetched = cursor.fetchall()

cursor.close()
db.close()
print(dataFetched)


data = []
for row in dataFetched:
    data.append(list(row))
    #print(list(row))
#print(data)

#按序号分组
DATA = [[] for i in range(100)]   #先分100列
for x in data:
    y = int(x[2])
    if y in range(1,92):
        DATA[y - 1].append(x)
    else:
        DATA[91].append(x)
    #print(x)


#新建新的表格
#使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
newwb = openpyxl.Workbook()

# 默认sheet
shet = newwb.active
shet.title = "temperature_field"

# shet.column_dismensions[‘A’].width = 40.0

#将数据按对应位置赋值到单元格内
font = Font('宋体')
fill_blue = PatternFill(fill_type ='solid',start_color='0000FF',end_color='000000')
fill_yellow = PatternFill(fill_type ='solid',start_color='FFFF00',end_color='000000')
fill_red = PatternFill(fill_type ='solid',start_color='FF0000',end_color='000000')
fill_no = PatternFill(fill_type ='solid',start_color='FFFFFF')
for j in range(len(DATA)):
    for i in range(len(DATA[j])):
        zi = ''
        zi += "{: <3}{: <6}{: <6}{: <3}{: <4}{: <3}".format(str(DATA[j][i][2]),
                                                                        str(DATA[j][i][3]),str(DATA[j][i][4]),str(int(DATA[j][i][6])),
                                                                        str(int(DATA[j][i][7])),str(int(DATA[j][i][8])))
        zi += DATA[j][i][5].strftime('%Y/%m/%d %H:%M:%S')
        shet.cell(i+1, j+1).value=zi
        shet.cell(i+1, j+1).font = font
        if DATA[j][i][7]==0 and DATA[j][i][8]==0:
            shet.cell(i + 1, j + 1).fill = fill_yellow
        elif DATA[j][i][7]==5 and DATA[j][i][8]==50:
            shet.cell(i + 1, j + 1).fill = fill_blue
        elif 0<DATA[j][i][3]<70 and 0<DATA[j][i][4]<100 and 25<DATA[j][i][6]<35 and  \
           0< DATA[j][i][7] < 100   and 10<DATA[j][i][8] <100:
            shet.cell(i + 1, j + 1).fill = fill_no
        else:
            shet.cell(i + 1, j + 1).fill = fill_red

        #print(zi)
#调整列宽
cols_list = list(shet.columns)
for i in range(0, len(cols_list)):
    if i >25:
        letter = chr(i // 26 + 64) + chr(i % 26 + 65)
    else:
        letter = chr(i % 26 + 65)
    shet.column_dimensions[letter].width = 45.0




path = 'newexl.xlsx'
#保存
# while True:
#     try:
#         newwb.save(path)
#         break;
#     except Exception as e:
#         print(str(e))
#         a = open(path, "w")
#         a.close()

try:
    newwb.save(path)

except Exception as e:
    print(str(e))
print('successful convert!')

