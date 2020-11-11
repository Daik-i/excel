
import wx
import os
import time
import pymysql
import openpyxl
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection
from datetime import datetime


def save_shuju():
    global path
    # 打开数据库连接
    while True:
        try:
            db = pymysql.connect(IP, name,mima, ku_name)
            break;
        except Exception as e:
            print(str(e))
            time.sleep(2)

    cursor = db.cursor()

    sql_str_read = "select * from temperature_field WHERE  submission_date between "+' " '+starttime+' " '+" and "+' " '+endtime+' " '+";"

    cursor.execute(sql_str_read)
    dataFetched = cursor.fetchall()


    cursor.close()
    db.close()
    #print(dataFetched)


    data = []
    for row in dataFetched:
        data.append(list(row))
        #print(list(row))
    #print(data)

    #按序号分组
    DATA = [[] for i in range(100)]   #先分100列
    for x in data:
        y = int(x[2])
        if y in range(1,92):
            DATA[y - 1].append(x)
        else:
            DATA[91].append(x)
        #print(x)


    #新建新的表格
    #使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
    newwb = openpyxl.Workbook()

    # 默认sheet
    shet = newwb.active
    shet.title = "temperature_field"

    # shet.column_dismensions[‘A’].width = 40.0

    #将数据按对应位置赋值到单元格内
    font = Font('宋体')
    fill_blue = PatternFill(fill_type ='solid',start_color='0000FF',end_color='000000')
    fill_yellow = PatternFill(fill_type ='solid',start_color='FFFF00',end_color='000000')
    fill_red = PatternFill(fill_type ='solid',start_color='FF0000',end_color='000000')
    fill_no = PatternFill(fill_type ='solid',start_color='FFFFFF')
    for j in range(len(DATA)):
        for i in range(len(DATA[j])):
            zi = ''
            zi += "{: <3}{: <6}{: <6}{: <3}{: <4}{: <3}".format(str(DATA[j][i][2]),
                                                                            str(DATA[j][i][3]),str(DATA[j][i][4]),str(int(DATA[j][i][6])),
                                                                            str(int(DATA[j][i][7])),str(int(DATA[j][i][8])))
            zi += DATA[j][i][5].strftime('%Y/%m/%d %H:%M:%S')
            shet.cell(i+1, j+1).value=zi
            shet.cell(i+1, j+1).font = font
            if DATA[j][i][7]==0 and DATA[j][i][8]==0:
                shet.cell(i + 1, j + 1).fill = fill_yellow
            elif DATA[j][i][7]==5 and DATA[j][i][8]==50:
                shet.cell(i + 1, j + 1).fill = fill_blue
            elif 0<DATA[j][i][3]<70 and 0<DATA[j][i][4]<100 and 25<DATA[j][i][6]<35 and  \
               0< DATA[j][i][7] < 100   and 10<DATA[j][i][8] <100:
                shet.cell(i + 1, j + 1).fill = fill_no
            else:
                shet.cell(i + 1, j + 1).fill = fill_red

            #print(zi)
    #调整列宽
    cols_list = list(shet.columns)
    for i in range(0, len(cols_list)):
        if i >25:
            letter = chr(i // 26 + 64) + chr(i % 26 + 65)
        else:
            letter = chr(i % 26 + 65)
        shet.column_dimensions[letter].width = 45.0
    try:
        newwb.save(path)

    except Exception as e:
        print(str(e))



number_ComboBox = 0
path = ''
IP = ''
ku_name = ""
name = ""
mima = ""
starttime ='2020-8-23 9:57:10'
endtime = ''

print(time)
class ExamplePanel(wx.Panel):

    def __init__(self, parent):
        wx.Panel.__init__(self, parent)
        self.quote = wx.StaticText(self, label='请选择:', pos=(20, 10))

        # 这个多行的文本框只是用来记录并显示events，不要纠结之
        self.logger = wx.TextCtrl(self, pos=(400,20), size=(250,300),
                                  style=wx.TE_MULTILINE | wx.TE_READONLY)

        # 一个按钮
        self.button = wx.Button(self, label='下 载', pos=(200, 325))
        self.Bind(wx.EVT_BUTTON, self.Save, self.button)
        # 一个按钮
        self.button1 = wx.Button(self, label='确 认', pos=(50, 325))
        self.Bind(wx.EVT_BUTTON, self.Config, self.button1)
        # 一个按钮
        self.button2 = wx.Button(self, label='关 闭', pos=(350, 325))
        self.Bind(wx.EVT_BUTTON, self.OnExit, self.button2)

        # 一个ComboBox控件（下拉菜单）
        self.sampleList = ['C:\\Users\\HP\\Desktop\\daik\\未命名.xlsx', 'C:\\Users\\HP\\Desktop\\未命名.xlsx','other']

        self.edithear = wx.ComboBox(self, pos=(20, 150), size=(330, -1),
                                    choices=self.sampleList,
                                    style=wx.CB_DROPDOWN |wx.TE_PROCESS_ENTER)
        self.edithear.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox, self.edithear)

        # 一个ComboBox控件（下拉菜单）
        self.sampleList1 = ["raspberrypi", '192.168.1.101']

        self.edithear1 = wx.ComboBox(self, pos=(20, 30), size=(150, -1),
                                    choices=self.sampleList1,
                                    style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear1.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox1, self.edithear1)

        # 一个ComboBox控件（下拉菜单）预备增加数据库表格选项
        self.sampleList2 = ["data1","data2"]
        self.edithear2 = wx.ComboBox(self, pos=(200, 30), size=(150, -1),
                                    choices=self.sampleList2,
                                    style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear2.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox2, self.edithear2)

        #数据库访问身份密码
        self.editname = wx.TextCtrl(self, value="root",pos=(20, 70), size=(150, -1))
        self.editmima = wx.TextCtrl(self, value="123",pos=(200, 70), size=(150, -1))

        # 数据获取区间
        self.begin = wx.TextCtrl(self, value=starttime,pos=(20, 110), size=(150, -1))
        self.end = wx.TextCtrl(self, value=endtime,pos=(200, 110), size=(150, -1))



    #路径事件处理
    def EvtComboBox(self, event):
        global number_ComboBox
        global path
        path1 =event.GetString()
        if path1 in ['other']:
            dlg = wx.FileDialog(self, "Choose path", path1, "未命名", "*.xlsx", wx.FD_SAVE)
            if dlg.ShowModal() == wx.ID_OK:
                self.filename = dlg.GetFilename()
                self.dirname = dlg.GetDirectory()
                path1 = self.dirname + '\\' + self.filename
                dlg.Destroy()
                self.edithear.Append(path1)
                self.edithear.SetSelection(3 + number_ComboBox)
                path = path1
            else:
                number_ComboBox -= 1
        else:
            number_ComboBox -= 1
            path = path1
        number_ComboBox += 1
        self.logger.AppendText('已选路径：\n' + path + '\n')

    #IP事件处理
    def EvtComboBox1(self, event):
        global IP
        IP =event.GetString()
        self.logger.AppendText('已选IP：\n' + IP + '\n')

    #数据库选择处理
    def EvtComboBox2(self, event):
        global ku_name
        ku_name =event.GetString()
        self.logger.AppendText('已选库：\n' + ku_name + '\n')





    def Config(self,event):
        global path,IP,ku_name,name,mima,starttime,endtime
        path = self.edithear.GetValue()
        IP = self.edithear1.GetValue()
        self.edithear1.Append(IP)
        ku_name = self.edithear2.GetValue()
        self.edithear2.Append(ku_name)
        name = self.editname.GetValue()
        mima = self.editmima.GetValue()
        starttime = self.begin.GetValue()
        endtime1 = self.end.GetValue()
        if endtime1 in['']:
            endtime = str(datetime.now())[:19]
            self.end.write(endtime)
        else:
            endtime = endtime1

        self.logger.AppendText('已选IP：\n' +IP+'\n'+'已选库：\n' +ku_name+'\n' + '用户名：\n' + name + '\n'+'密码：\n' +mima+'\n' \
                                +'已选路径：\n' + path + '\n'+'起始时间：\n' +starttime+'\n' + '结束时间：\n' + endtime + '\n' )


    def Save(self,event):
        save_shuju()
        self.logger.AppendText('数据保存成功'+'\n')

    def OnExit(self, e):
        frame.Close(True)



app = wx.App(False)
frame = wx.Frame(None, size=(690,400),title ='数据库下载')
panel = ExamplePanel(frame)
frame.Show()
app.MainLoop()