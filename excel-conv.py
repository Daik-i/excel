import openpyxl
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection
#获取表格
wb = openpyxl.load_workbook('temperature_field.xlsx')

#获取页
sheet = wb['temperature_field']
#print(sheet)

# 获取所有数据
data_list = list(sheet.rows)[1:]  # 剔除第一条数据，第一条是标题
data = []  # 新列表储存数据#
for row in data_list:
    # 每一行的数据存储起来 每循环一次清空
    row_data = []
    # 把 row 里面的数据依次提取出来
    for cell in row:
        # 把值添加到 row_data 列表中
        row_data.append(cell.value)
    # 每循环一次，往 new_data 添加一次数据
    data.append(row_data)
print(data)
wb.close()

#按序号分组
DATA = [[] for i in range(100)]   #先分100列
for x in data:
    y = int(x[2])
    DATA[y-1].append(x)
#print(DATA)

#新建新的表格
#使用openpyxl没有必要先在系统中新建一个.xlsx，我们需要做的只需要引Workbook这个类，接着开始调用它。
newwb = openpyxl.Workbook()

# 默认sheet
shet = newwb.active
shet.title = "temperature_field"

#将数据按对应位置赋值到单元格内
for j in range(len(DATA)):
    font = Font('宋体')
    for i in range(len(DATA[j])):
        zi = ''
        zi += "{: <3}{: <6}{: <6}{: <3}{: <4}{: <3}".format(str(DATA[j][i][2]),
                                                                        str(DATA[j][i][3]),str(DATA[j][i][4]),str(DATA[j][i][6]),
                                                                        str(DATA[j][i][7]),str(DATA[j][i][8]))
        zi += DATA[j][i][5].strftime('%Y/%m/%d %H:%M:%S')
        shet.cell(i+1, j+1).value=zi
        shet.cell(i+1, j+1).font = font

        #print(zi)


#保存
newwb.save('newexl.xlsx')
print('successful convert!')