# -*- coding: utf-8 -*-
import os

from PIL import Image
import numpy as np
import openpyxl
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection

def conver_color(x):
    color_str = ''
    for i in x:
        if i > 15 :
            color_str += hex(i)[2:]
        else:
            color_str += '0'+hex(i)[2:]
    return color_str.upper()

MAX_WIDTH = 300
MAX_HEIGHT = 300

def resize(img):
    w, h = img.size
    if w > MAX_WIDTH:
        h = MAX_WIDTH / w * h
        w = MAX_WIDTH

    if h > MAX_HEIGHT:
        w = MAX_HEIGHT / h * w
        h = MAX_HEIGHT
    return img.resize((int(w), int(h)), Image.ANTIALIAS)

def loadImage():
    im = Image.open("lena.jpg")
    im = resize(im)
    size=im.size
    im.show()
    data = im.getdata()
    im.close()
    DATA = []
    for i in data:
        color = conver_color(i)
        DATA.append(color)
    matx = np.matrix(DATA)
    matx = np.reshape(matx ,size)

    newwb = openpyxl.Workbook()
    shet = newwb.active
    shet.title = "picture"

    for i in range(size[0]):
        for j in range(size[1]):
            shet.cell(i + 1, j + 1).fill = PatternFill(fill_type ='solid',fgColor=matx[i,j])
    cols_list = list(shet.columns)                              #获取列名
    for i in range(0, len(cols_list)):                          #将数字序列转化为excel格式的列名序列
        if i > 25:
            letter = chr(i // 26 + 64) + chr(i % 26 + 65)
        else:
            letter = chr(i % 26 + 65)
        shet.column_dimensions[letter].width = 0.25                #每列列宽5

    rows_list = list(shet.rows)                                 #获取列名

    for i in range(0, len(rows_list)):                          #将数字序列转化为excel格式的列名序列
        shet.row_dimensions[i+1].height = 1.5                #每列列宽5

    #保存表格
    try:
        newwb.save("pic.xlsx")
        newwb.close()
        print('successful convert!')
    except Exception as e:
        print(str(e))
        print('5')

path = "pic.xlsx"

loadImage()