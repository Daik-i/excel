# -*- coding: utf-8 -*-
import os
import wx
from PIL import Image
import numpy as np
import openpyxl
from openpyxl.styles import fills, colors, NamedStyle, Font, Side, Border, PatternFill, Alignment, Protection

#将像素信息从（**，**，**）转化为openpyxl可识别的字符串
def conver_color(x):
    color_str = ''
    for i in x:
        if i > 15:
            color_str += hex(i)[2:]
        else:
            color_str += '0' + hex(i)[2:]
    return color_str.upper()


MAX_WIDTH = 300
MAX_HEIGHT = 300

#将图片尺寸缩小，像素过多会导致excel报错，一般300×300没问题
def resize(img):
    w, h = img.size
    if w > MAX_WIDTH:
        h = MAX_WIDTH / w * h
        w = MAX_WIDTH

    if h > MAX_HEIGHT:
        w = MAX_HEIGHT / h * w
        h = MAX_HEIGHT
    return img.resize((int(w), int(h)), Image.ANTIALIAS)


def loadImage():
    im = Image.open(sources)
    im = resize(im)
    size = im.size
    # print(size)
    # im.show()
    data = im.getdata()             #信息提取
    im.close()
    DATA = []
    for i in data:
        color = conver_color(i)
        DATA.append(color)
    matx = np.matrix(DATA)
    matx = np.reshape(matx, (size[1],size[0]))     #将字符串重新整合为矩阵形式，矩阵行列刚好和图片尺寸相反

    newwb = openpyxl.Workbook()
    shet = newwb.active
    shet.title = "picture"

    for i in range(size[1]):
        for j in range(size[0]):
            shet.cell(i + 1, j + 10).fill = PatternFill(fill_type='solid', fgColor=matx[i, j])
    cols_list = list(shet.columns)  # 获取列名
    for i in range(9, len(cols_list)):  # 将数字序列转化为excel格式的列名序列
        if i > 25:
            letter = chr(i // 26 + 64) + chr(i % 26 + 65)
        else:
            letter = chr(i % 26 + 65)
        shet.column_dimensions[letter].width = 0.25  # 每列列宽0.25

    rows_list = list(shet.rows)  # 获取列名

    for i in range(0, len(rows_list)):  # 将数字序列转化为excel格式的行序列
        shet.row_dimensions[i + 1].height = 1.5  # 每行行高1.5

    # 保存表格
    try:
        newwb.save(path)
        newwb.close()
        print('successful convert!')
    except Exception as e:
        print(str(e))
path = "pic.xlsx"
sources= "pic.jpg"
number_ComboBox = 0
number_ComboBox1 = 0
class ExamplePanel(wx.Panel):

    def __init__(self, parent):
        wx.Panel.__init__(self, parent)
        self.quote = wx.StaticText(self, label='请选择文件:', pos=(20, 10))
        self.quote1 = wx.StaticText(self, label='请选择路径:', pos=(20, 70))
        # 这个多行的文本框只是用来记录并显示events，不要纠结之
        self.logger = wx.TextCtrl(self, pos=(400, 20), size=(250, 300),
                                  style=wx.TE_MULTILINE | wx.TE_READONLY)
        # 一个按钮
        self.button = wx.Button(self, label='转 换', pos=(200, 325))
        self.Bind(wx.EVT_BUTTON, self.Save, self.button)
        # 一个按钮
        self.button1 = wx.Button(self, label='确 认', pos=(50, 325))
        self.Bind(wx.EVT_BUTTON, self.Config, self.button1)
        # 打开按钮
        self.button3 = wx.Button(self, label='打 开', pos=(500, 325))
        self.Bind(wx.EVT_BUTTON, self.OnOpen, self.button3)
        # 一个按钮
        self.button2 = wx.Button(self, label='关 闭', pos=(350, 325))
        self.Bind(wx.EVT_BUTTON, self.OnExit, self.button2)

        # 一个ComboBox控件（下拉菜单）
        self.sampleList = ["", 'other']

        self.edithear = wx.ComboBox(self, pos=(20, 90), size=(330, -1),
                                    choices=self.sampleList,
                                    style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox, self.edithear)

        self.sampleList1 = ['C:\\Users\\HP\\Desktop\\daik\\picture.jpg', 'C:\\Users\\HP\\Desktop\\picture.jpg',
                           'other']
        self.edithear1 = wx.ComboBox(self, pos=(20,30), size=(330, -1),
                                     choices=self.sampleList1,
                                     style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear1.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox1, self.edithear1)

    def EvtComboBox1(self, event):
        global number_ComboBox1
        global sources
        sources1 = event.GetString()
        if sources1 in ['other']:
            dlg = wx.FileDialog(self, "Choose file", sources, "picture", "*.jpg", wx.FD_OPEN)
            if dlg.ShowModal() == wx.ID_OK:
                self.filename = dlg.GetFilename()
                self.dirname = dlg.GetDirectory()
                sources1 = self.dirname + '\\' + self.filename
                dlg.Destroy()
                self.edithear1.Append(sources1)
                self.edithear1.SetSelection(3 + number_ComboBox1)
                sources = sources1
            else:
                number_ComboBox1 -= 1
        else:
            number_ComboBox1 -= 1
            sources = sources1
        number_ComboBox1 += 1
        self.logger.AppendText('已选文件：\n' + sources + '\n')

    # 路径事件处理
    def EvtComboBox(self, event):
        global number_ComboBox
        global path
        path1 = event.GetString()
        if path1 in ['other']:
            dlg = wx.FileDialog(self, "Choose path", path1, "未命名", "*.xlsx", wx.FD_SAVE)
            if dlg.ShowModal() == wx.ID_OK:
                self.filename = dlg.GetFilename()
                self.dirname = dlg.GetDirectory()
                path1 = self.dirname + '\\' + self.filename
                dlg.Destroy()
                self.edithear.Append(path1)
                self.edithear.SetSelection(2 + number_ComboBox)
                path = path1
            else:
                number_ComboBox -= 1
        else:
            number_ComboBox -= 1
            path = path1
        number_ComboBox += 1
        self.logger.AppendText('已选路径：\n' + path + '\n')


    def Config(self, event):
        global path,sources
        sources = self.edithear1.GetValue()
        path = self.edithear.GetValue()
        if path in['']:
            path = sources[:-4]+".xlsx"
            self.edithear.Append(path)
        self.logger.AppendText('已选文件：\n' + sources + '\n')
        self.logger.AppendText('已选路径：\n' + path + '\n')

    def OnOpen(self,event):
        global path
        os.startfile(path)


    def Save(self, event):
        loadImage()
        self.logger.AppendText('图片转换成功' + '\n')

    def OnExit(self, e):
        frame.Close(True)


app = wx.App(False)
frame = wx.Frame(None, size=(690, 400), title='图片转换')
panel = ExamplePanel(frame)
frame.Show()
app.MainLoop()
