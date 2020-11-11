# -*- coding: utf-8 -*-
import wx
import os
import time
import pymysql
import openpyxl
from openpyxl.styles import fills,colors,NamedStyle,Font,Side,Border,PatternFill,Alignment,Protection
from datetime import datetime

#较前一版增加选前段和后段的功能，增加了打开功能，自动前推一天的功能

#执行下载
def save_shuju():
    global path,num1,num2
    # 打开数据库连接
    while True:
        try:
            db = pymysql.connect(IP, name,mima, ku_name)
            break;
        except Exception as e:
            print(str(e))
            time.sleep(2)
    cursor = db.cursor()
    sql_str_read = "select * from temperature_field WHERE  submission_date between "+' " '+starttime+' " '+" and "+' " '+endtime+' " '+";"
    cursor.execute(sql_str_read)
    dataFetched = cursor.fetchall()
    cursor.close()
    db.close()
    #print(dataFetched)

    #将元组数据转化为列表数据
    data1 = []
    for row in dataFetched:
        data1.append(list(row))

    #对前段和后段的选择进行操作，找到选取段的端点
    if num2 !=0:
        if num2<len(data1):
            point1 = len(data1) - num2
        else:
            point1 =0
        point2 = len(data1)-1
    elif num1 !=0:
        if num1<len(data1):
            point2 = num1-1
        else:
            point2 = len(data1)-1
        point1 = 0
    else:
        point1 = 0                      #起始点
        point2 = len(data1) - 1         #终止点



    # 按传感器号分组
    DATA = [[] for i in range(100)]   #先分100列

    for x in data1[point1:point2]:
        y = int(x[2])
        if y in range(1,92):
            DATA[y - 1].append(x)
        else:
            DATA[91].append(x)
        #print(x)

    #新建新的表格
    newwb = openpyxl.Workbook()

    # 默认sheet
    shet = newwb.active
    shet.title = "temperature_field"

    #提前设置单元格颜色格式
    font = Font('宋体')
    fill_blue = PatternFill(fill_type ='solid',start_color='0000FF',end_color='000000')
    fill_yellow = PatternFill(fill_type ='solid',start_color='FFFF00',end_color='000000')
    fill_red = PatternFill(fill_type ='solid',start_color='FF0000',end_color='000000')
    fill_no = PatternFill(fill_type ='solid',start_color='FFFFFF')
    # 将数据按对应位置赋值到单元格内
    for j in range(len(DATA)):
        for i in range(len(DATA[j])):
            zi = ''
            zi += "{: <3}{: <6}{: <6}{: <3}{: <4}{: <3}".format(str(DATA[j][i][2]),
                                                                            str(DATA[j][i][3]),str(DATA[j][i][4]),str(int(DATA[j][i][6])),
                                                                            str(int(DATA[j][i][7])),str(int(DATA[j][i][8])))
            zi += DATA[j][i][5].strftime('%Y/%m/%d %H:%M:%S')
            shet.cell(i+1, j+1).value=zi                                                        #单元格赋值
            shet.cell(i+1, j+1).font = font                                                     #设置字体
            if DATA[j][i][7]==0 and DATA[j][i][8]==0:                                           #出现双零设为黄色
                shet.cell(i + 1, j + 1).fill = fill_yellow
            elif DATA[j][i][7]==5 and DATA[j][i][8]==50:                                        #初始状态设为蓝色
                shet.cell(i + 1, j + 1).fill = fill_blue
            elif 0<DATA[j][i][3]<70 and 0<DATA[j][i][4]<100 and 25<DATA[j][i][6]<35 and  \
               0< DATA[j][i][7] < 100   and 10<DATA[j][i][8] <100:                              #正常范围为无填充颜色
                shet.cell(i + 1, j + 1).fill = fill_no
            else:                                                                               #错误数据为红色
                shet.cell(i + 1, j + 1).fill = fill_red

    #调整列宽
    cols_list = list(shet.columns)                              #获取列名
    for i in range(0, len(cols_list)):                          #将数字序列转化为excel格式的列名序列
        if i >25:
            letter = chr(i // 26 + 64) + chr(i % 26 + 65)
        else:
            letter = chr(i % 26 + 65)
        shet.column_dimensions[letter].width = 45.0             #每列列宽45

    #保存表格
    try:
        newwb.save(path)
    except Exception as e:
        print(str(e))

#所需变量，大多不需要提前赋值
number_ComboBox = 0         #用于路径的下拉菜单
path = ''                   #路径
IP = ''                     #IP地址
ku_name = ""                #数据库名称
name = ""                   #用户名
mima = ""                   #密码
#设置默认开始时间，一天前
if str(datetime.now())[8:10] in ['01']:
    if str(datetime.now())[5:7] in ['05','07','10','12']:
        starttime = str(datetime.now())[0:5] + str(int(str(datetime.now())[5:7]) - 1)+'-30' + str(datetime.now())[10:19]
    elif str(datetime.now())[5:7] in ['02','04','06','08','09','11']:
        starttime = str(datetime.now())[0:5] + str(int(str(datetime.now())[5:7]) - 1)+'-31' + str(datetime.now())[10:19]
    elif str(datetime.now())[5:7] in ['03']:
        if str(datetime.now())[0:4]%4==0:
            starttime = str(datetime.now())[0:5] + str(int(str(datetime.now())[5:7])-1)+'-29'+str(datetime.now())[10:19]
        else:
            starttime = str(datetime.now())[0:5]+str(int(str(datetime.now())[5:7])-1)+'-28'+str(datetime.now())[10:19]
    else:
        starttime = str(int(str(datetime.now())[0:4])-1) +'-12-31'+str(datetime.now())[10:19]
else :
    starttime = str(datetime.now())[0:8]+str(int(str(datetime.now())[8:10])-1)+str(datetime.now())[10:19]

endtime = ''                    #截至时间，默认现在
num1,num2 =0,0                  #前段，后段数量设置的变量

#GUI面板部分
class ExamplePanel(wx.Panel):
    def __init__(self, parent):
        wx.Panel.__init__(self, parent)
        # 提示文字
        self.quote = wx.StaticText(self, label='请选择:', pos=(20, 10))

        # 这个多行的文本框只是用来记录并显示events，#多行，只读文本框
        self.logger = wx.TextCtrl(self, pos=(400,20), size=(250,300),
                                  style=wx.TE_MULTILINE | wx.TE_READONLY)

        # 下载按钮
        self.button = wx.Button(self, label='下 载', pos=(200, 325))
        self.Bind(wx.EVT_BUTTON, self.Save, self.button)
        # 确认按钮
        self.button1 = wx.Button(self, label='确 认', pos=(50, 325))
        self.Bind(wx.EVT_BUTTON, self.Config, self.button1)
        # 打开按钮
        self.button3 = wx.Button(self, label='打 开', pos=(500, 325))
        self.Bind(wx.EVT_BUTTON, self.OnOpen, self.button3)
        # 关闭按钮
        self.button2 = wx.Button(self, label='关 闭', pos=(350, 325))
        self.Bind(wx.EVT_BUTTON, self.OnExit, self.button2)


        # 路径下拉菜单
        self.sampleList = ['C:\\Users\\HP\\Desktop\\daik\\未命名.xlsx', 'C:\\Users\\HP\\Desktop\\未命名.xlsx','other']
        self.edithear = wx.ComboBox(self, pos=(20, 190), size=(330, -1),
                                    choices=self.sampleList,
                                    style=wx.CB_DROPDOWN |wx.TE_PROCESS_ENTER)
        self.edithear.SetSelection(0)                                       #默认第一项
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox, self.edithear)         #事件绑定

        # IP地址下拉菜单
        self.sampleList1 = ["raspberrypi", '192.168.1.101']
        self.edithear1 = wx.ComboBox(self, pos=(20, 30), size=(150, -1),
                                    choices=self.sampleList1,
                                    style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear1.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox1, self.edithear1)

        # 数据库下拉菜单（预备增加数据库表格选项）
        self.sampleList2 = ["data1","data2"]
        self.edithear2 = wx.ComboBox(self, pos=(200, 30), size=(150, -1),
                                    choices=self.sampleList2,
                                    style=wx.CB_DROPDOWN | wx.TE_PROCESS_ENTER)
        self.edithear2.SetSelection(0)
        self.Bind(wx.EVT_COMBOBOX, self.EvtComboBox2, self.edithear2)

        #数据库访问身份/密码
        self.editname = wx.TextCtrl(self, value="root",pos=(20, 70), size=(150, -1))
        self.editmima = wx.TextCtrl(self, value="123",pos=(200, 70), size=(150, -1))

        # 数据时间区间
        self.begin = wx.TextCtrl(self, value=starttime,pos=(20, 110), size=(150, -1))
        self.end = wx.TextCtrl(self, value=endtime,pos=(200, 110), size=(150, -1))

        # 数据数量区间
        self.num_begin = wx.TextCtrl(self, value=str(num1), pos=(20, 150), size=(150, -1))
        self.num_end = wx.TextCtrl(self, value=str(num2), pos=(200, 150), size=(150, -1))

        # self.scroller = wx.ScrolledWindow(self, -1)
        # self.scroller.SetScrollbars(1, 1, 1440, 900)

    #路径事件处理
    def EvtComboBox(self, event):
        global number_ComboBox
        global path
        path1 =event.GetString()
        if path1 in ['other']:          #如果选择第三项则跳出路径选择框
            dlg = wx.FileDialog(self, "请选择路径：", path1, "未命名", "*.xlsx", wx.FD_SAVE)    #打开路径选择框
            if dlg.ShowModal() == wx.ID_OK:         #如果确认就保存路径，有效下拉事件加1，用来在选择other项后文本框能够显示选择内容
                self.filename = dlg.GetFilename()
                self.dirname = dlg.GetDirectory()
                path1 = self.dirname + '\\' + self.filename
                dlg.Destroy()
                self.edithear.Append(path1)
                self.edithear.SetSelection(3 + number_ComboBox)
                path = path1
                number_ComboBox += 1
        else:
            path = path1
        self.logger.AppendText('已选路径：\n' + path + '\n')

    #IP事件处理
    def EvtComboBox1(self, event):
        global IP
        IP =event.GetString()
        self.logger.AppendText('已选IP：\n' + IP + '\n')

    #数据库选择处理
    def EvtComboBox2(self, event):
        global ku_name
        ku_name =event.GetString()
        self.logger.AppendText('已选库：\n' + ku_name + '\n')

    # 打开文件
    def OnOpen(self, event):
        global path
        os.startfile(path)

    #文本框手输内容的采集
    def Config(self,event):
        global path,IP,ku_name,name,mima,starttime,endtime, num1,num2
        path = self.edithear.GetValue()
        IP = self.edithear1.GetValue()
        self.edithear1.Append(IP)
        ku_name = self.edithear2.GetValue()
        self.edithear2.Append(ku_name)
        name = self.editname.GetValue()
        mima = self.editmima.GetValue()
        starttime = self.begin.GetValue()
        endtime1 = self.end.GetValue()
        if endtime1 in['']:
            endtime = str(datetime.now())[:19]
            self.end.write(endtime)
        else:
            endtime = endtime1
        num1 =int(self.num_begin.GetValue())
        num2 =int(self.num_end.GetValue())
        self.logger.AppendText('已选IP：\n' +IP+'\n'+'已选库：\n' +ku_name+'\n' + '用户名：\n' + name + '\n'+'密码：\n' +mima+'\n' \
                                +'已选路径：\n' + path + '\n'+'起始时间：\n' +starttime+'\n' + '结束时间：\n' + endtime + '\n' \
                                 '前段：\n' + str(num1) + '\n' + '后段：\n' + str(num2) + '\n' )

    #下载并保存
    def Save(self,event):
        save_shuju()
        self.logger.AppendText('数据保存成功'+'\n')
    #退出程序
    def OnExit(self, e):
        frame.Close(True)
#GUI开始运行
app = wx.App(False)
frame = wx.Frame(None, size=(690,400),title ='数据库下载')
panel = ExamplePanel(frame)
frame.Show()
app.MainLoop()